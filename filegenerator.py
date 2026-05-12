import copy
import csv
import logging
import os
import pathlib
import shutil
import struct
import subprocess
import tempfile
from pathlib import Path
from tempfile import NamedTemporaryFile

import pdfkit
from django.conf import settings
from django.contrib.staticfiles.finders import find as find_static_file
from django.http import HttpResponse, StreamingHttpResponse
from django.template import Context, Template
from django.template.loader import get_template
from django.utils.text import get_valid_filename
from openpyxl import Workbook

from mighty.apps import MightyConfig as conf
from mighty.functions import make_searchable

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Shared PDF margin helpers — used by both generate_pdf and PDFView so they
# produce identical top/bottom margins for the same header/footer HTML.
# ---------------------------------------------------------------------------

def get_wkhtmltoimage_bin() -> str | None:
    path = shutil.which('wkhtmltoimage')
    if path:
        return path
    wkpdf = shutil.which('wkhtmltopdf')
    if wkpdf:
        sibling = Path(wkpdf).with_name('wkhtmltoimage')
        if sibling.is_file():
            return str(sibling)
    return None


def read_png_pixel_height(png_path: Path) -> int:
    with png_path.open('rb') as f:
        if f.read(8) != b'\x89PNG\r\n\x1a\n':
            raise ValueError('not a PNG')
        while True:
            chunk_head = f.read(8)
            if len(chunk_head) < 8:
                raise ValueError('truncated PNG')
            chunk_len, chunk_type = struct.unpack('>I4s', chunk_head)
            chunk_data = f.read(chunk_len)
            f.read(4)
            if chunk_type == b'IHDR':
                _, height = struct.unpack('>II', chunk_data[:8])
                return height


def a4_page_width_mm(orientation: str) -> float:
    if (orientation or 'Portrait').lower() == 'landscape':
        return 297.0
    return 210.0


def pdf_screen_dpi(options: dict | None = None) -> int:
    raw = (options or {}).get('dpi')
    if raw in (None, ''):
        return 96
    try:
        return int(raw)
    except (TypeError, ValueError):
        return 96


def page_width_px_for_pdf(options: dict | None, orientation: str) -> int:
    mm = a4_page_width_mm(orientation)
    dpi = pdf_screen_dpi(options)
    return max(1, int(round(mm / 25.4 * dpi)))


def run_wkhtmltoimage(
    bin_path: str,
    source_html_path: Path,
    png_path: Path,
    width_px: int,
) -> bool:
    html_abs = str(source_html_path.resolve())
    png_abs = str(png_path.resolve())
    base = [
        bin_path, '--quiet', '--javascript-delay', '200',
        '--width', str(width_px), html_abs, png_abs,
    ]
    for with_local_access in (True, False):
        cmd = list(base)
        if with_local_access:
            cmd.insert(1, '--enable-local-file-access')
        try:
            subprocess.run(cmd, check=True, capture_output=True, timeout=45, text=True)
        except (FileNotFoundError, subprocess.TimeoutExpired) as exc:
            logger.warning('wkhtmltoimage failed: %s', exc)
            return False
        except subprocess.CalledProcessError as exc:
            logger.debug(
                'wkhtmltoimage exit %s (local=%s): %s',
                exc.returncode, with_local_access,
                (exc.stderr or exc.stdout or '')[:800],
            )
            continue
        else:
            return True
    return False


def auto_margin_top_from_header(
    header_html_path: str,
    orientation: str,
    options: dict | None = None,
    config: dict | None = None,
) -> str | None:
    """Return ``margin-top`` sized to the rendered header height.

    Returns ``None`` on failure so the caller keeps its default margin.
    Optional config keys: ``header_margin_padding_mm`` (default 3),
    ``header_margin_top_min_in`` (default 0.35),
    ``header_margin_top_max_in`` (default 3.0).
    Pass ``config={'auto_header_margin': False}`` to skip probing.
    """
    config = config or {}
    if config.get('auto_header_margin') is False:
        return None
    bin_path = get_wkhtmltoimage_bin()
    if not bin_path:
        return None
    dpi = pdf_screen_dpi(options)
    width_px = page_width_px_for_pdf(options, orientation)
    fd, png_name = tempfile.mkstemp(suffix='.png')
    os.close(fd)
    png_path = Path(png_name)
    try:
        if not run_wkhtmltoimage(bin_path, Path(header_html_path), png_path, width_px):
            return None
        height_px = read_png_pixel_height(png_path)
    except ValueError:
        logger.warning('Header height probe returned an invalid PNG')
        return None
    finally:
        png_path.unlink(missing_ok=True)

    if height_px < 8:
        return None
    padding_mm = float(config.get('header_margin_padding_mm', 3))
    height_in = (height_px / dpi) + (padding_mm / 25.4)
    min_in = float(config.get('header_margin_top_min_in', 0.35))
    max_in = float(config.get('header_margin_top_max_in', 3.0))
    height_in = max(min_in, min(max_in, height_in))
    return f'{height_in:.3f}in'


def auto_margin_bottom_from_footer(
    footer_html_path: str,
    orientation: str,
    options: dict | None = None,
    config: dict | None = None,
) -> str | None:
    """Return ``margin-bottom`` sized to the rendered footer height.

    Returns ``None`` on failure so the caller keeps its default margin.
    Optional config keys: ``footer_margin_padding_mm`` (default 0),
    ``footer_margin_bottom_min_in`` (default 0.25),
    ``footer_margin_bottom_max_in`` (default 3.0).
    Pass ``config={'auto_footer_margin': False}`` to skip probing.
    """
    config = config or {}
    if config.get('auto_footer_margin') is False:
        return None
    bin_path = get_wkhtmltoimage_bin()
    if not bin_path:
        return None
    dpi = pdf_screen_dpi(options)
    width_px = page_width_px_for_pdf(options, orientation)
    fd, png_name = tempfile.mkstemp(suffix='.png')
    os.close(fd)
    png_path = Path(png_name)
    try:
        if not run_wkhtmltoimage(bin_path, Path(footer_html_path), png_path, width_px):
            return None
        height_px = read_png_pixel_height(png_path)
    except ValueError:
        logger.warning('Footer height probe returned an invalid PNG')
        return None
    finally:
        png_path.unlink(missing_ok=True)

    if height_px < 8:
        return None
    padding_mm = float(config.get('footer_margin_padding_mm', 0))
    height_in = (height_px / dpi) + (padding_mm / 25.4)
    min_in = float(config.get('footer_margin_bottom_min_in', 0.25))
    max_in = float(config.get('footer_margin_bottom_max_in', 3.0))
    height_in = max(min_in, min(max_in, height_in))
    return f'{height_in:.3f}in'


class StreamingBuffer:
    def write(self, value):
        return value


def _build_pdf_font_face_style():
    fonts = {}
    fonts.update(getattr(settings, 'PDFMAKER', {}).get('fonts', {}))
    fonts.update(getattr(settings, 'FONTS_FILES_WEASYPRINT', {}))

    rules = []
    for font_file, font_name in fonts.items():
        font_path = find_static_file(f'fonts/{font_file}')
        if not font_path:
            continue
        rules.extend([
            '@font-face {',
            f"  font-family: '{font_name}';",
            f"  src: url('file://{font_path}') format('truetype');",
            '  font-style: normal;',
            '  font-weight: 400;',
            '}',
        ])

    if not rules:
        return ''
    return '<style>' + '\n'.join(rules) + '</style>'


def _inject_font_style_in_head(html):
    style = _build_pdf_font_face_style()
    if not style:
        return html
    if '</head>' in html:
        return html.replace('</head>', f'{style}</head>', 1)
    return style + html


class FileGenerator:
    filename = None
    items = []
    fields = ()
    ct_list = {
        'text/csv': 'csv',
        'application/vnd.ms-excel': 'xls',
        'application/vnd.ms-excel': 'excel',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': 'xlsx',
        'application/pdf': 'pdf',
        'text/html': 'html',
        'application/json': 'json',
        'application/xml': 'xml',
        'text/plain': 'txt',
    }
    row = 0
    col = 0

    @property
    def reverse_ct_list(self):
        return {v: k for k, v in self.ct_list.items()}

    def __init__(self, *args, **kwargs):
        for name, data in kwargs.items():
            setattr(self, name, data)

    @property
    def iter_rows(self):
        yield self.fields
        yield from self.items

    def get_by_content_type(self, ct='text/csv', response='http'):
        return getattr(self, f'response_{response}')(self.ct_list[ct])

    def get_by_format(self, ct='csv', response='http'):
        return getattr(self, f'response_{response}')(ct)

    def get_filename(self, ct):
        return f'{get_valid_filename(make_searchable(self.filename))}.{ct}'

    def response_file(self, ct):
        items_method = f'iter_items_{ct}'
        getattr(self, items_method)(
            self.items, open(self.get_filename(ct), 'w', encoding='utf-8')
        )

    # EXCEL
    def iter_items_xls(self, items, ws):
        for row in self.iter_rows:
            ws.append(row)

    def file_xlsx(self, ext, ct):
        return self.file_xls(ext, ct)

    def file_xls(self, ext, ct):
        wb = Workbook()
        ws = wb.active
        self.iter_items_xls(self.items, ws)
        tmp = open(self.get_filename(ext), 'rb')
        wb.save(tmp.name)
        return tmp

    def http_excel(self, ext, ct):
        return self.http_xls('xlsx', ct)

    def http_xlsx(self, ext, ct):
        return self.http_xls(ext, ct)

    def http_xls(self, ext, ct):
        wb = Workbook()
        ws = wb.active
        self.iter_items_xls(self.items, ws)
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        response = HttpResponse(content=stream, content_type=ct)
        response['Content-Disposition'] = (
            'attachment;filename=' + self.get_filename(ext)
        )
        return response

    # CSV
    def iter_stream_csv(self, items, pseudo_buffer):
        writer = csv.writer(pseudo_buffer, quoting=csv.QUOTE_ALL)
        for item in self.iter_rows:
            yield writer.writerow(item)

    def iter_items_csv(self, items, pseudo_buffer, stream=True):
        writer = csv.writer(pseudo_buffer, quoting=csv.QUOTE_ALL)
        for item in self.iter_rows:
            writer.writerow(item)

    def file_csv(self, ext, ct=None):
        tmp = open(self.get_filename(ext), 'w', encoding='utf-8')
        self.iter_items_csv(self.items, tmp, False)
        tmp.close()
        return open(self.get_filename(ext), 'rb')

    def http_csv(self, ext, ct):
        response = StreamingHttpResponse(
            streaming_content=self.iter_stream_csv(
                self.items, StreamingBuffer()
            ),
            content_type=ct,
        )
        response['Content-Disposition'] = (
            'attachment;filename=' + self.get_filename(ext)
        )
        return response

    # PDF
    def iter_items_pdf(self, items):
        headers = ''.join(f'<th>{field}</th>' for field in self.fields)
        rows = ''.join(
            f'<tr>{"".join(f"<td>{i if i is not None else ''}</td>" for i in item)}</tr>'
            for item in items
        )
        return f'<table><thead><tr>{headers}</tr></thead><tbody>{rows}</tbody></table>'

    def generate_pdf_from_html(self, output_path=None):
        from datetime import datetime

        import pdfkit

        html_string = Template(self.html).render(
            Context({
                'table': self.iter_items_pdf(self.items),
                'items': self.items,
                'document_name': self.filename,
                'queryset': self.queryset,
            })
        )
        html_string = _inject_font_style_in_head(html_string)
        # Date du jour formatée
        today = datetime.today().strftime('%d/%m/%Y')

        # Texte de pied de page avec pagination
        footer_text = (
            f'Octolo certifié en date du {today} - Page [page] sur [topage]'
        )

        # Options wkhtmltopdf
        options = {
            'page-size': 'A4',
            'orientation': 'Landscape',
            'margin-top': '20mm',
            'margin-bottom': '20mm',
            'margin-left': '20mm',
            'margin-right': '20mm',
            'encoding': 'UTF-8',
            'enable-local-file-access': None,
            'footer-center': footer_text,
            'footer-font-size': '10',
            'footer-line': '',  # Ligne de séparation (optionnel)
        }
        return pdfkit.from_string(html_string, output_path, options=options)

    def file_pdf(self, ext, ct):
        return self.generate_pdf_from_html(output_path=self.get_filename('pdf'))

    def http_pdf(self, ext, ct):
        pdf_bytes = self.generate_pdf_from_html(False)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'inline; filename={self.get_filename("pdf")}'
        )
        return response

    def response_file(self, ct):
        return getattr(self, 'file_' + ct)(ct, self.reverse_ct_list[ct])

    def response_http(self, ct):
        return getattr(self, 'http_' + ct)(ct, self.reverse_ct_list[ct])


def generate_pdf(**kwargs):
    _header_enable, _footer_enable = False, False
    file_name = kwargs.get('file_name', False)
    # Deep copy so we never mutate the shared conf.pdf_options dict across calls.
    options = copy.deepcopy(kwargs.get('options', conf.pdf_options))
    header = kwargs.get('header', False)
    footer = kwargs.get('footer', False)
    context = kwargs.get('context', {})
    content = kwargs.get('content', False)
    content_html = kwargs.get('content_html', False)
    as_string = kwargs.get('as_string', False)
    orientation = options.get('orientation', 'Portrait')
    context.update({
        'static': os.path.abspath(settings.STATIC_ROOT),
    })

    # header
    if header:
        header_tpl = kwargs.get(
            'conf_header',
            get_template(conf.pdf_header).render({'header': header}),
        )
        context.update({'header_html': header})
        header_html = tempfile.NamedTemporaryFile(suffix='.html', delete=False)
        header_html.write(
            Template(header_tpl).render(Context(context)).encode('utf-8')
        )
        header_html.close()
        header = header_html
        options['--header-html'] = header_html.name
        auto_top = auto_margin_top_from_header(
            header_html.name, orientation, options
        )
        if auto_top:
            options['margin-top'] = auto_top
    else:
        options['margin-top'] = '0.30in'

    if footer:
        footer_tpl = kwargs.get(
            'conf_footer',
            get_template(conf.pdf_footer).render({'footer': footer}),
        )
        context.update({'footer_html': footer})
        footer_html = tempfile.NamedTemporaryFile(suffix='.html', delete=False)
        footer_html.write(
            Template(footer_tpl).render(Context(context)).encode('utf-8')
        )
        footer_html.close()
        footer = footer_html
        options['--footer-html'] = footer_html.name
        auto_bottom = auto_margin_bottom_from_footer(
            footer_html.name, orientation, options
        )
        if auto_bottom:
            options['margin-bottom'] = auto_bottom
    else:
        options['margin-bottom'] = '0.30in'

    if (content or content_html) and file_name:
        tmp_pdf = tempfile.NamedTemporaryFile(suffix='.pdf', delete=True)
        if content_html:
            content_html = Template(content_html).render(Context(context))
        else:
            content_html = get_template(content).render(context)
        content_tpl = kwargs.get(
            'conf_content',
            get_template(conf.pdf_content).render({
                'content': content_html,
                'body_lead': '<br/>' if header else '',
            }),
        )
        content_html = Template(content_tpl).render(Context(context))
        content_html = _inject_font_style_in_head(content_html)
        options['enable-local-file-access'] = None
        pdfkit.from_string(content_html, tmp_pdf.name, options=options)
        path_tmp = tmp_pdf.name
        valid_file_name = get_valid_filename(file_name)
        # Use a unique per-call temp directory to avoid collisions on a fixed
        # path like /tmp/<file_name> (which fails with PermissionError on
        # rootless/distroless containers when the file already exists).
        final_pdf = str(pathlib.Path(tempfile.mkdtemp()) / valid_file_name)
        shutil.copyfile(path_tmp, final_pdf)

    # a verifier
    if footer:
        pathlib.Path(footer_html.name).unlink()
    if header:
        pathlib.Path(header_html.name).unlink()
    if as_string:
        tmp_pdf.close()
        pathlib.Path(final_pdf).unlink()
        return content_html
    return final_pdf, tmp_pdf
