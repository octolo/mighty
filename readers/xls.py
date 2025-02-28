import contextlib

from openpyxl import load_workbook


def reader(iterator):
    total = [[col.value for col in row] for row in iterator]
    yield from total


class ReaderXLS:
    def __init__(self, filename, worksheet='', *args, **kwargs):
        self.wb = load_workbook(filename, **kwargs)
        self.ws = (worksheet and self.wb[worksheet]) or self.wb.active
        self.total = self.ws.max_row
        self.reader = reader(self.ws)
        self._fieldnames = kwargs.get('fieldnames')
        self.restkey = kwargs.get('restkey')
        self.restval = kwargs.get('restval')
        self.skip_blank_lines = kwargs.get('skip_blank_lines', False)
        self.line_num = 0

    @property
    def fieldnames(self):
        if self._fieldnames is None:
            with contextlib.suppress(StopIteration):
                self._fieldnames = next(self.reader)
        self.line_num += 1
        return self._fieldnames

    @fieldnames.setter
    def fieldnames(self, value):
        self._fieldnames = value

    def __iter__(self):
        return self

    def __next__(self):
        if self.line_num == 0:
            self.fieldnames
        row = next(self.reader)
        self.line_num += 1

        while self.skip_blank_lines and all(cell is None for cell in row):
            row = next(self.reader)

        d = dict(zip(self.fieldnames, row, strict=False))
        lf = len(self.fieldnames)
        lr = len(row)
        if lf < lr:
            d[self.restkey] = row[lf:]
        elif lf > lr:
            for key in self.fieldnames[lr:]:
                d[key] = self.restval
        return d
